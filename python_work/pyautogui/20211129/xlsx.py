from openpyxl import Workbook,load_workbook

class Customer_Email:
    def __init__(self):
        pass

    def SaveEmail(self, email):
        self.e = email
        wb = load_workbook("Customer_Mail.xlsx")
        ws = wb.active

        ws.append([self.e])

        wb.save("Customer_Mail.xlsx")
        wb.close()

    def CreateEmail_xl(self):
        wb = Workbook() # 새 워크북 생성
        ws = wb.active # 현재 활성화된 sheet 가져옴
        ws.title = "emal" # sheet 의 이름을 변경
        wb.save("Customer_Mail.xlsx")
        wb.close()