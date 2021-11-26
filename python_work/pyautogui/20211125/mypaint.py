import pyautogui
import time
import pyperclip

from openpyxl import Workbook

def fontcopy(str):
    pyperclip.copy(str)

def doms():
    try:
        time.sleep(1)
        pyautogui.hotkey('win','r')
        pyautogui.write('mspaint')
        pyautogui.hotkey('enter')
        time.sleep(1)
        box = pyautogui.locateOnScreen('font.PNG', confidence=0.8)
        pyautogui.click(box)
        pyautogui.move(10, 200, duration=1)
        pyautogui.click()
        fontcopy('참 잘했어요')
        pyautogui.hotkey('ctrl','v')

        mywindow = pyautogui.getActiveWindow()
        mywindow.close()

        time.sleep(1)
        pyautogui.write('n')

        wb = Workbook()
        ws = wb.active

        count = ws['A1'].value
        if count == None:
            count = '1'
        else:
            count = int(count)+1

        ws['A1'] = count
        ws['B1'] = count+' idid'

        wb.save('log.xlsx')
        wb.close()
        
    except Exception as e:
        print(e)














