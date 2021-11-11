from openpyxl import *

'''
    Workbook() 엑셀 파일이 없을때 생성
    load_workbook('a.xlsx') 엑셀파일이 있을때 불러오는것

    wb.save('a.xlsx')
    wb.close()
'''

class MExcel:
    def __init__(self):
        pass

    def save(self,a,b,c):
        wb = load_workbook('mexcel.xlsx')
        ws = wb.active
        ws.append([a,b,c])
        wb.save('mexcel.xlsx')
        wb.close()

    def load(self):
        wb = load_workbook('mexcel.xlsx')
        ws = wb.active

        # iter_rows 행을 기준으로 반복하는 함수
        # iter_cols 열을 기준으로 반복하는 함수
        for row in ws.iter_rows():
            for cell in row:  # 11 22 33
                print(cell.value, end=' ')
            print()

        wb.save('mexcel.xlsx')
        wb.close()

    def create(self):
        wb = Workbook()
        wb.save('mexcel.xlsx')
        wb.close()

