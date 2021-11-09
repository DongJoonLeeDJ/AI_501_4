from openpyxl import Workbook
from random import *

'''
    d.xlsx 파일 만들어서
    국어 영어 수학 넣고
    랜덤으로 0~100까지 만들어 10개 행 넣기
'''
def dxlsx(filename):
    wb = Workbook()
    ws = wb.active

    ws.append(["국어","영어","수학"])
    for i in range(1,11):
        ws.append([randint(0,100),randint(0,100),randint(0,100)])

    colran = ws["B:C"]

    for col in colran:
        for cell in col:
            print(cell.value)
    print()

    for column in ws.iter_cols():
        print("column ",end="")
        print(column)

    for row in ws.iter_rows():
        print("row ", end="")
        print(row)

    for row in range(1,ws.max_row+1):
        for col in range(1, ws.max_column + 1):
            print(ws.cell(column=col,row=row).value)

    wb.save(filename+".xlsx")
    wb.close()
