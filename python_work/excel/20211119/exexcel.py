from openpyxl import *

''' 파일 생성 하고 기본 데이터 넣는 함수'''
def makebasic():
    wb = Workbook()
    ws = wb.active

    scores = [
        (1, 10, 8, 5, 14, 26, 12),
        (2, 7, 3, 7, 15, 24, 18),
        (3, 9, 5, 8, 8, 12, 4),
        (4, 7, 8, 7, 17, 21, 18),
        (5, 7, 8, 7, 16, 25, 15),
        (6, 3, 5, 8, 8, 17, 0),
        (7, 4, 9, 10, 16, 27, 18),
        (8, 6, 6, 6, 15, 19, 17),
        (9, 10, 10, 9, 19, 30, 19),
        (10, 9, 8, 8, 20, 25, 20)
    ]

    ws.append(['학번', '출석', '퀴즈1', '퀴즈2', '중간고사', '기말고사', '프로젝트'])
    for s in scores:  # 기존 성적 데이터 넣기
        ws.append(s)

    wb.save('score.xlsx')
    wb.close()


def dmajum():
    wb = load_workbook('score.xlsx')
    ws = wb.active

    for idx,cell in enumerate(ws['D']):
        if idx ==0:
            continue
        else:
            cell.value = 10


    wb.save('score.xlsx')
    wb.close()
    
    
def 총점함수():
    wb = load_workbook('score.xlsx')
    ws = wb.active

    ws['H1'] = "총점"

    for row in ws.iter_rows(min_row=2,min_col=2):
        sum = 0
        try:
            for cell in row[0:6]:
                sum += cell.value
            print('sum',sum)
        except Exception as e:
            print(e)
        row[6].value = sum
        # print(row[0].value,
        #       row[1].value,
        #       row[2].value,
        #       row[3].value,
        #       row[4].value,
        #       row[5].value,
        #       row[6].value,
        #       row[7].value,
        #       )

    wb.save('score.xlsx')
    wb.close()
    
def 성적함수():
    wb = load_workbook('score.xlsx')
    ws = wb.active

    ws['I1'] = "성적"

    wb.save('score.xlsx')
    wb.close()


















